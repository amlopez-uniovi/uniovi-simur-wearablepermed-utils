import pytest
from uniovi_simur_wearablepermed_utils.file_management import *
import numpy as np
import os
from datetime import datetime, date, time, timedelta
import openpyxl
import pandas as pd

def test_read_time_from_excel_valid_time():
    # Create a mock Excel file with a valid time
    file_path = 'tests/data_import/mock_excel.xlsx'
    sheet_name = 'Sheet1'
    cell_reference = 'A1'
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet[cell_reference] = '12:34:56'
    workbook.save(file_path)
    
    result = read_time_from_excel(file_path, sheet_name, cell_reference)
    expected = time(12, 34, 56)
    
    assert result == expected, f"Expected {expected}, but got {result}"

def test_read_time_from_excel_invalid_format():
    # Create a mock Excel file with an invalid time format
    file_path = 'tests/data_import/mock_excel.xlsx'
    sheet_name = 'Sheet1'
    cell_reference = 'A1'
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet[cell_reference] = 'invalid_time'
    workbook.save(file_path)
    
    result = read_time_from_excel(file_path, sheet_name, cell_reference)
    
    assert result is None, f"Expected None, but got {result}"
    


def test_read_time_from_excel_timedelta():
    # Create a mock Excel file with a timedelta value
    file_path = 'tests/data_import/mock_excel.xlsx'
    sheet_name = 'Sheet1'
    cell_reference = 'A1'
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet[cell_reference] = timedelta(hours=5, minutes=30, seconds=15)
    workbook.save(file_path)
    
    result = read_time_from_excel(file_path, sheet_name, cell_reference)
    expected = time(5, 30, 15)
    
    assert result == expected, f"Expected {expected}, but got {result}"

def test_read_time_from_excel_numeric():
    # Create a mock Excel file with a numeric value representing time
    file_path = 'tests/data_import/mock_excel.xlsx'
    sheet_name = 'Sheet1'
    cell_reference = 'A1'
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet[cell_reference] = 0.5  # Represents 12:00:00 PM
    workbook.save(file_path)
    
    result = read_time_from_excel(file_path, sheet_name, cell_reference)
    expected = time(12, 0, 0)
    
    assert result == expected, f"Expected {expected}, but got {result}"

def test_read_time_from_excel_time_object():
    # Create a mock Excel file with a time object
    file_path = 'tests/data_import/mock_excel.xlsx'
    sheet_name = 'Sheet1'
    cell_reference = 'A1'
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet[cell_reference] = time(14, 45, 30)
    workbook.save(file_path)
    
    result = read_time_from_excel(file_path, sheet_name, cell_reference)
    expected = time(14, 45, 30)
    
    assert result == expected, f"Expected {expected}, but got {result}"
    
def test_read_date_from_excel_valid_date():
    # Create a mock Excel file with a valid date
    file_path = 'tests/data_import/mock_excel.xlsx'
    sheet_name = 'Sheet1'
    cell_reference = 'A1'
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet[cell_reference] = '15/07/2024'
    workbook.save(file_path)
    
    result = read_date_from_excel(file_path, sheet_name, cell_reference)
    expected = date(2024, 7, 15)
    
    assert result == expected, f"Expected {expected}, but got {result}"

def test_read_date_from_excel_invalid_format():
    # Create a mock Excel file with an invalid date format
    file_path = 'tests/data_import/mock_excel.xlsx'
    sheet_name = 'Sheet1'
    cell_reference = 'A1'
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet[cell_reference] = 'invalid_date'
    workbook.save(file_path)
    
    result = read_date_from_excel(file_path, sheet_name, cell_reference)
    
    assert result is None, f"Expected None, but got {result}"


def test_read_date_from_excel_numeric():
    # Create a mock Excel file with a numeric value representing a date
    file_path = 'tests/data_import/mock_excel.xlsx'
    sheet_name = 'Sheet1'
    cell_reference = 'A1'
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet[cell_reference] = 44757  # Represents 15/07/2022 in Excel date format
    workbook.save(file_path)
    
    result = read_date_from_excel(file_path, sheet_name, cell_reference)
    expected = date(2022, 7, 15)
    
    assert result == expected, f"Expected {expected}, but got {result}"


def test_extract_WPM_info_from_excel():
    expected_output = {
        "Hora de inicio de acelerómetro muslo (hh:mm:ss) - Hora de ordenador": time(9, 41, 6),
        "Hora de inicio de acelerómetro cadera (hh:mm:ss) - Hora de ordenador": time(9, 35, 51),
        "Hora de inicio de acelerómetro muñeca (hh:mm:ss) - Hora de ordenador": time(9, 35, 21),
        "Fecha día 1": datetime(2024, 7, 8,0, 0),
        "FASE REPOSO CON K5 - Hora de inicio": time(9, 47, 45),
        "FASE REPOSO CON K5 - Hora de fin": time(10, 22, 45),
        "TAPIZ RODANTE - Hora de inicio": time(10, 45, 5),
        "TAPIZ RODANTE - Hora de fin": time(11, 13, 5),
        "SIT TO STAND 30 s - Hora de inicio": time(11, 31, 19),
        "SIT TO STAND 30 s - Hora de fin": time(11, 31, 49),
        "INCREMENTAL CICLOERGOMETRO - Hora de inicio REPOSO": time(11, 51, 35),
        "INCREMENTAL CICLOERGOMETRO - Hora de inicio CALENTAMIENTO": time(11, 54, 35),
        "INCREMENTAL CICLOERGOMETRO - Hora de inicio INCREMENTAL": time(11, 57, 35),
        "INCREMENTAL CICLOERGOMETRO - Hora de fin": time(12, 7, 18),
        "Fecha día 7": datetime(2024, 7, 15,0,0),
        "YOGA - Hora de inicio": time(12, 36, 32),
        "YOGA - Hora de fin": time(12, 38, 32),
        "SENTADO VIENDO LA TV - Hora de inicio": time(12, 39, 0),
        "SENTADO VIENDO LA TV - Hora de fin": time(12, 41, 0),
        "SENTADO LEYENDO - Hora de inicio": time(12, 41, 16),
        "SENTADO LEYENDO - Hora de fin": time(12, 43, 16),
        "SENTADO USANDO PC - Hora de inicio": time(12, 44, 46),
        "SENTADO USANDO PC - Hora de fin": time(12, 46, 46),
        "DE PIE USANDO PC - Hora de inicio": time(12, 47, 13),
        "DE PIE USANDO PC - Hora de fin": time(12, 49, 13),
        "DE PIE DOBLANDO TOALLAS - Hora de inicio": time(12, 51, 57),
        "DE PIE DOBLANDO TOALLAS - Hora de fin": time(12, 53, 57),
        "DE PIE MOVIENDO LIBROS - Hora de inicio": time(12, 54, 46),
        "DE PIE MOVIENDO LIBROS - Hora de fin": time(12, 56, 46),
        "DE PIE BARRIENDO - Hora de inicio": time(12, 57, 14),
        "DE PIE BARRIENDO - Hora de fin": time(12, 59, 14),
        "CAMINAR USUAL SPEED - Hora de inicio": time(13, 1, 30),
        "CAMINAR USUAL SPEED - Hora de fin": time(13, 3, 30),
        "CAMINAR CON MÓVIL O LIBRO - Hora de inicio": time(13, 3, 48),
        "CAMINAR CON MÓVIL O LIBRO - Hora de fin": time(13, 5, 48),
        "CAMINAR CON LA COMPRA - Hora de inicio": time(13, 6, 14),
        "CAMINAR CON LA COMPRA - Hora de fin": time(13, 8, 14),
        "CAMINAR ZIGZAG - Hora de inicio": time(13, 8, 55),
        "CAMINAR ZIGZAG - Hora de fin": time(13, 10, 55),
        "TROTAR - Hora de inicio": time(13, 11, 16),
        "TROTAR - Hora de fin": time(13, 13, 16),
        "SUBIR Y BAJAR ESCALERAS - Hora de inicio": time(13, 15, 16),
        "SUBIR Y BAJAR ESCALERAS - Hora de fin": time(13, 17, 16),
        'ACTIVIDAD NO ESTRUCTURADA - Hora de inicio': time(12, 30),
        'ACTIVIDAD NO ESTRUCTURADA - Hora de fin': time(12, 15)
    }
    
    result = extract_WPM_info_from_excel('tests/data_import/PMPXXX_RegistroActividades.xlsx')
    assert result == expected_output
  

def test_load_MATRIX_data_by_index():
    """Test loading MATRIX data by index with correct axis transformations."""
 
    # Expected output after applying axis transformations
    expected_output = np.array([[1, -2, 4, -3, -5, 7, -6, -10, -20, -30, -40],
                                [10, -20, 40, -30, -50, 70, -60, -1, -2, -3, -4]])

    # Call function (you'll need to replace with actual data loading function or mock)
    loaded_data = load_MATRIX_data_by_index('tests/data_import/MATA00_few_data.csv', np.array([-1, 3, -2]))
    
    loaded_data = loaded_data*1
    assert np.allclose(loaded_data, expected_output)
    def test_load_WPM_data_wrist():
        # Create a mock CSV file with sample data for the wrist
        csv_file = 'tests/data_import/mock_wrist_data.csv'
        data = {
            'dateTime': [1625097600, 1625097601, 1625097602],  # Example timestamps
            'acc_x': [0.1, 0.2, 0.3],
            'acc_y': [0.4, 0.5, 0.6],
            'acc_z': [0.7, 0.8, 0.9],
            'gyr_x': [1.0, 1.1, 1.2],
            'gyr_y': [1.3, 1.4, 1.5],
            'gyr_z': [1.6, 1.7, 1.8],
            'bodySurface_temp': [36.5, 36.6, 36.7],
            'ambient_temp': [22.0, 22.1, 22.2],
            'hr_raw': [70, 71, 72],
            'hr': [75, 76, 77]
        }
        df = pd.DataFrame(data)
        df.to_csv(csv_file, index=False)

        # Expected output after applying axis transformations
        expected_output = np.array([
            [1625097600, -0.1, 0.7, -0.4, -1.0, 1.6, -1.3, 36.5, 22.0, 70, 75],
            [1625097601, -0.2, 0.8, -0.5, -1.1, 1.7, -1.4, 36.6, 22.1, 71, 76],
            [1625097602, -0.3, 0.9, -0.6, -1.2, 1.8, -1.5, 36.7, 22.2, 72, 77]
        ])

        # Call the function
        result = load_WPM_data(csv_file, 'Wrist')

        # Check if the result matches the expected output
        assert np.allclose(result, expected_output), f"Expected {expected_output}, but got {result}"


def test_load_WPM_data_thigh():
    # Create a mock CSV file with sample data for the thigh
    csv_file = 'tests/data_import/mock_thigh_data.csv'
    data = {
        'dateTime': [1625097600, 1625097601, 1625097602],  # Example timestamps
        'acc_x': [0.1, 0.2, 0.3],
        'acc_y': [0.4, 0.5, 0.6],
        'acc_z': [0.7, 0.8, 0.9],
        'gyr_x': [1.0, 1.1, 1.2],
        'gyr_y': [1.3, 1.4, 1.5],
        'gyr_z': [1.6, 1.7, 1.8],
        'bodySurface_temp': [36.5, 36.6, 36.7],
        'ambient_temp': [22.0, 22.1, 22.2],
        'hr_raw': [70, 71, 72],
        'hr': [75, 76, 77]
    }
    df = pd.DataFrame(data)
    df.to_csv(csv_file, index=False)

    # Expected output after applying axis transformations
    expected_output = np.array([
        [1625097600, 0.7, -0.1, 0.4, 1.6, -1.0, 1.3, 36.5, 22.0, 70, 75],
        [1625097601, 0.8, -0.2, 0.5, 1.7, -1.1, 1.4, 36.6, 22.1, 71, 76],
        [1625097602, 0.9, -0.3, 0.6, 1.8, -1.2, 1.5, 36.7, 22.2, 72, 77]
    ])

    # Call the function
    result = load_WPM_data(csv_file, 'Thigh')

    # Check if the result matches the expected output
    assert np.allclose(result, expected_output), f"Expected {expected_output}, but got {result}"


def test_load_WPM_data_hip():
    # Create a mock CSV file with sample data for the hip
    csv_file = 'tests/data_import/mock_hip_data.csv'
    data = {
        'dateTime': [1625097600, 1625097601, 1625097602],  # Example timestamps
        'acc_x': [0.1, 0.2, 0.3],
        'acc_y': [0.4, 0.5, 0.6],
        'acc_z': [0.7, 0.8, 0.9],
        'gyr_x': [1.0, 1.1, 1.2],
        'gyr_y': [1.3, 1.4, 1.5],
        'gyr_z': [1.6, 1.7, 1.8],
        'bodySurface_temp': [36.5, 36.6, 36.7],
        'ambient_temp': [22.0, 22.1, 22.2],
        'hr_raw': [70, 71, 72],
        'hr': [75, 76, 77]
    }
    df = pd.DataFrame(data)
    df.to_csv(csv_file, index=False)

    # Expected output after applying axis transformations
    expected_output = np.array([
        [1625097600, -0.1, -0.7, -0.4, -1.0, -1.6, -1.3, 36.5, 22.0, 70, 75],
        [1625097601, -0.2, -0.8, -0.5, -1.1, -1.7, -1.4, 36.6, 22.1, 71, 76],
        [1625097602, -0.3, -0.9, -0.6, -1.2, -1.8, -1.5, 36.7, 22.2, 72, 77]
    ])

    # Call the function
    result = load_WPM_data(csv_file, 'Hip')

    # Check if the result matches the expected output
    assert np.allclose(result, expected_output), f"Expected {expected_output}, but got {result}"


def test_calculate_accelerometer_drift_no_walk_start_sample():
    # Create a mock Excel file with power-on and power-off dates and times
    excel_file_path = 'tests/data_import/mock_activity_log.xlsx'
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Hoja1'
    
    # Power-on date and time
    sheet['E13'] = '08/07/2024'
    sheet['E37'] = '00:00:00'
    
    # Power-off date and time
    sheet['E112'] = '09/07/2024'
    sheet['E273'] = '00:00:00'
    
    workbook.save(excel_file_path)
    
    # Mock WPM data
    WPM_data = np.array([
        [1720389600, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        [1720476000, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0] 
    ])
    
    # Call the function
    K = calculate_accelerometer_drift(WPM_data, excel_file_path, 'Thigh')
    
    # Expected scaling factor
    expected_K = 1.0
    
    assert K*1000 == expected_K, f"Expected {expected_K}, but got {K}"

def test_calculate_accelerometer_drift_with_walk_start_sample():
    # Create a mock Excel file with power-on date and time, and walk start date and time
    excel_file_path = 'tests/data_import/mock_activity_log.xlsx'
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Hoja1'
    
    # Power-on date and time
    sheet['E13'] = '08/07/2024'
    sheet['E37'] = '00:00:00'
    
    # Walk start date and time
    sheet['E112'] = '14/07/2024'
    sheet['D219'] = '00:00:00'
    
    workbook.save(excel_file_path)
    
    # Mock WPM data
    WPM_data = np.array([
        [1720389600, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        [1720908000, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]  
    ])
    
    # Call the function
    K = calculate_accelerometer_drift(WPM_data, excel_file_path, 'Thigh', walk_usual_speed_start_sample=1)
    
    # Expected scaling factor
    expected_K = 1.0
    
    assert K*1000 == expected_K, f"Expected {expected_K}, but got {K}"

def test_calculate_accelerometer_drift_insufficient_data():
    # Create a mock Excel file with only power-on date and time
    excel_file_path = 'tests/data_import/mock_activity_log.xlsx'
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Hoja1'
    
    # Power-on date and time
    sheet['E13'] = '08/07/2024'
    sheet['E37'] = '09:41:06'
    
    workbook.save(excel_file_path)
    
    # Mock WPM data
    WPM_data = np.array([
        [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        [604800000, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]  # 7 days in milliseconds
    ])
    
    # Call the function
    K = calculate_accelerometer_drift(WPM_data, excel_file_path, 'Thigh')
    
    # Expected scaling factor
    expected_K = 1.0
    
    assert K == expected_K, f"Expected {expected_K}, but got {K}"

def test_apply_scaling_to_matrix_data():
    """Test the apply_scaling_to_matrix_data function with a known scaling factor."""
    # Mock data with timestamps (in milliseconds) and some arbitrary values
    original_data = np.array([
        [1000, 1, 2, 3],
        [2000, 4, 5, 6],
        [3000, 7, 8, 9],
        [4000, 10, 11, 12],
        [5000, 13, 14, 15]
    ])

    # Known scaling factor
    K = 2.0

    # Expected scaled data
    expected_scaled_data = np.array([
        [1000, 1, 2, 3],
        [1500, 4, 5, 6],
        [2000, 7, 8, 9],
        [2500, 10, 11, 12],
        [3000, 13, 14, 15]
    ])

    # Call the function
    scaled_data = apply_scaling_to_matrix_data(original_data, K)

    # Check if the scaled data matches the expected data
    assert np.allclose(scaled_data, expected_scaled_data), "The scaled data does not match the expected data."

def test_apply_scaling_to_matrix_data_no_scaling():
    """Test the apply_scaling_to_matrix_data function with a scaling factor of 1 (no scaling)."""
    # Mock data with timestamps (in milliseconds) and some arbitrary values
    original_data = np.array([
        [1000, 1, 2, 3],
        [2000, 4, 5, 6],
        [3000, 7, 8, 9],
        [4000, 10, 11, 12],
        [5000, 13, 14, 15]
    ])

    # Scaling factor of 1 (no scaling)
    K = 1.0

    # Expected scaled data (should be the same as original data)
    expected_scaled_data = original_data.copy()

    # Call the function
    scaled_data = apply_scaling_to_matrix_data(original_data, K)

    # Check if the scaled data matches the expected data
    assert np.allclose(scaled_data, expected_scaled_data), "The scaled data does not match the expected data when no scaling is applied."

def test_apply_scaling_to_matrix_data_negative_scaling():
    """Test the apply_scaling_to_matrix_data function with a negative scaling factor."""
    # Mock data with timestamps (in milliseconds) and some arbitrary values
    original_data = np.array([
        [1000, 1, 2, 3],
        [2000, 4, 5, 6],
        [3000, 7, 8, 9],
        [4000, 10, 11, 12],
        [5000, 13, 14, 15]
    ])

    # Negative scaling factor
    K = -2.0

    # Expected scaled data
    expected_scaled_data = np.array([
        [1000, 1, 2, 3],
        [500, 4, 5, 6],
        [0, 7, 8, 9],
        [-500, 10, 11, 12],
        [-1000, 13, 14, 15]
    ])

    # Call the function
    scaled_data = apply_scaling_to_matrix_data(original_data, K)

    # Check if the scaled data matches the expected data
    assert np.allclose(scaled_data, expected_scaled_data), "The scaled data does not match the expected data with negative scaling factor."
